from django.contrib import admin
from simple_history.admin import SimpleHistoryAdmin
# Register your models here.
from core.models import *
from modeltranslation.admin import TranslationAdmin


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

admin.site.register(Setting)
admin.site.register(Category)

admin.site.register(Contact)
admin.site.register(Subscriber)
admin.site.register(Brands)
admin.site.register(Size)
admin.site.register(Tag)
admin.site.register(Color)


@admin.register(News)
class NewsAdmin(TranslationAdmin):
    list_display = ('title', 'created_at', 'updated_at', 'like', 'dislike', 'views',)
    search_fields = ('title', 'content',)
    list_filter = ('created_at', 'updated_at')
    ordering = ('created_at',)
    readonly_fields = ('like', 'dislike', 'views', 'created_at', 'updated_at',)
    fieldsets = (
        (None, {
            'fields': ('title', 'content','category', 'image','is_active')
        }),
        ('Advanced options', {
            'classes': ('collapse',),
            'fields': ('like', 'dislike', 'views', 'created_at', 'updated_at',)
        }),
    )


def export_to_excel(modeladmin, request, queryset):
    
    wb = Workbook()
    ws = wb.active
    
   
    headers = ['Title', 'Content', 'Price', 'Discount', 'Category']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws.column_dimensions[col_letter].width = 15
    
    
    for row_num, product in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=product.title)
        ws.cell(row=row_num, column=2, value=product.content)
        ws.cell(row=row_num, column=3, value=product.price)
        ws.cell(row=row_num, column=4, value=product.discount)
        ws.cell(row=row_num, column=5, value=str(product.category))
    
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)
    return response


export_to_excel.short_description = "Export to Excel"



@admin.register(Product)
class ProductAdmin(SimpleHistoryAdmin):
    list_display = ('title', 'created_at', 'updated_at', 'price', 'discount', 'brands', 'size', 'color', 'tags', 'category')
    search_fields = ('title', 'content',)
    list_filter = ('created_at', 'updated_at', 'category')
    ordering = ('created_at',)
    readonly_fields = ('created_at', 'updated_at',)
    fieldsets = (
        (None, {
            'fields': ('title', 'content', 'image', 'price', 'discount', 'brands', 'size', 'color', 'tags', 'category', 'slug')
        }),
        ('Translations', {
            'fields': ('title_en', 'content_en', 'title_tr', 'content_tr', 'title_az', 'content_az'),  # Include your translation fields here
            'classes': ('collapse',)  
        }),
    )
    actions = [export_to_excel]
    
    
    
    
    def save_model(self, request, obj, form, change):
        if change and form.initial['discount'] != form.cleaned_data['discount']:
            
            changed_by = request.user
            
            print(f"The discount of product {obj.title} has been changed by {changed_by}.")

        super().save_model(request, obj, form, change)
    